from datetime import datetime, timedelta
from flask import send_file
import io
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ReportGenerator:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_LEFT
        )
    
    def generate_pdf_report(self, agendamentos, pacientes, profissionais, servicos, report_type="geral"):
        """Gera relatório em PDF"""
        buffer = io.BytesIO()
        
        # Cria o documento PDF
        if report_type == "agendamentos":
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = self._create_appointments_pdf(agendamentos, pacientes, profissionais, servicos)
        elif report_type == "financeiro":
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = self._create_financial_pdf(agendamentos, profissionais)
        else:
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = self._create_general_pdf(agendamentos, pacientes, profissionais, servicos)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _create_appointments_pdf(self, agendamentos, pacientes, profissionais, servicos):
        """Cria PDF de agendamentos"""
        story = []
        
        # Título
        title = Paragraph("Relatório de Agendamentos", self.title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Informações gerais
        info_text = f"Período: {datetime.now().strftime('%d/%m/%Y')} - Total de agendamentos: {len(agendamentos)}"
        info_para = Paragraph(info_text, self.styles['Normal'])
        story.append(info_para)
        story.append(Spacer(1, 20))
        
        # Tabela de agendamentos
        if agendamentos:
            data = [['Data', 'Hora', 'Paciente', 'Profissional', 'Serviço', 'Status', 'Valor']]
            
            for ag in agendamentos:
                paciente = next((p for p in pacientes if p.id == ag.paciente_id), None)
                profissional = next((p for p in profissionais if p.id == ag.profissional_id), None)
                servico = next((s for s in servicos if s.id == ag.servico_id), None)
                
                data.append([
                    ag.data,
                    ag.hora,
                    paciente.nome if paciente else 'N/A',
                    profissional.nome if profissional else 'N/A',
                    servico.nome if servico else 'N/A',
                    ag.status,
                    f"R$ {ag.valor_pago:.2f}" if ag.valor_pago else 'R$ 0,00'
                ])
            
            table = Table(data, colWidths=[1*inch, 0.8*inch, 1.5*inch, 1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        return story
    
    def _create_financial_pdf(self, agendamentos, profissionais):
        """Cria PDF financeiro"""
        story = []
        
        # Título
        title = Paragraph("Relatório Financeiro", self.title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Resumo financeiro
        total_receita = sum(ag.valor_pago for ag in agendamentos if ag.valor_pago)
        agendamentos_pagos = [ag for ag in agendamentos if ag.valor_pago and ag.valor_pago > 0]
        
        summary_data = [
            ['Total de Agendamentos', len(agendamentos)],
            ['Agendamentos Pagos', len(agendamentos_pagos)],
            ['Total de Receita', f"R$ {total_receita:.2f}"],
            ['Receita Média por Agendamento', f"R$ {(total_receita/len(agendamentos_pagos) if agendamentos_pagos else 0):.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Receita por profissional
        prof_receita = {}
        for ag in agendamentos:
            if ag.valor_pago and ag.valor_pago > 0:
                prof_id = ag.profissional_id
                if prof_id not in prof_receita:
                    prof_receita[prof_id] = 0
                prof_receita[prof_id] += ag.valor_pago
        
        if prof_receita:
            prof_data = [['Profissional', 'Receita Total']]
            for prof_id, receita in prof_receita.items():
                profissional = next((p for p in profissionais if p.id == prof_id), None)
                prof_data.append([
                    profissional.nome if profissional else f'Profissional {prof_id}',
                    f"R$ {receita:.2f}"
                ])
            
            prof_table = Table(prof_data, colWidths=[3*inch, 2*inch])
            prof_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(prof_table)
        
        return story
    
    def _create_general_pdf(self, agendamentos, pacientes, profissionais, servicos):
        """Cria PDF geral com estatísticas"""
        story = []
        
        # Título
        title = Paragraph("Relatório Geral da Clínica", self.title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Estatísticas gerais
        total_pacientes = len(pacientes)
        total_profissionais = len(profissionais)
        total_servicos = len(servicos)
        
        stats_data = [
            ['Total de Pacientes', total_pacientes],
            ['Total de Profissionais', total_profissionais],
            ['Total de Serviços', total_servicos],
            ['Total de Agendamentos', len(agendamentos)]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightcoral),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Status dos agendamentos
        status_count = {}
        for ag in agendamentos:
            status = ag.status
            status_count[status] = status_count.get(status, 0) + 1
        
        if status_count:
            status_data = [['Status', 'Quantidade']]
            for status, count in status_count.items():
                status_data.append([status, str(count)])
            
            status_table = Table(status_data, colWidths=[3*inch, 2*inch])
            status_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightyellow),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(status_table)
        
        return story
    
    def generate_excel_report(self, agendamentos, pacientes, profissionais, servicos, report_type="geral"):
        """Gera relatório em Excel"""
        wb = openpyxl.Workbook()
        
        if report_type == "agendamentos":
            self._create_appointments_excel(wb, agendamentos, pacientes, profissionais, servicos)
        elif report_type == "financeiro":
            self._create_financial_excel(wb, agendamentos, profissionais)
        else:
            self._create_general_excel(wb, agendamentos, pacientes, profissionais, servicos)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_appointments_excel(self, wb, agendamentos, pacientes, profissionais, servicos):
        """Cria planilha Excel de agendamentos"""
        ws = wb.active
        ws.title = "Agendamentos"
        
        # Cabeçalhos
        headers = ['Data', 'Hora', 'Paciente', 'Profissional', 'Serviço', 'Status', 'Valor', 'Observações']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados
        for row, ag in enumerate(agendamentos, 2):
            paciente = next((p for p in pacientes if p.id == ag.paciente_id), None)
            profissional = next((p for p in profissionais if p.id == ag.profissional_id), None)
            servico = next((s for s in servicos if s.id == ag.servico_id), None)
            
            ws.cell(row=row, column=1, value=ag.data)
            ws.cell(row=row, column=2, value=ag.hora)
            ws.cell(row=row, column=3, value=paciente.nome if paciente else 'N/A')
            ws.cell(row=row, column=4, value=profissional.nome if profissional else 'N/A')
            ws.cell(row=row, column=5, value=servico.nome if servico else 'N/A')
            ws.cell(row=row, column=6, value=ag.status)
            ws.cell(row=row, column=7, value=ag.valor_pago or 0)
            ws.cell(row=row, column=8, value=ag.observacoes or '')
        
        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_financial_excel(self, wb, agendamentos, profissionais):
        """Cria planilha Excel financeira"""
        ws = wb.active
        ws.title = "Financeiro"
        
        # Resumo
        ws.cell(row=1, column=1, value="Relatório Financeiro").font = Font(bold=True, size=16)
        ws.cell(row=3, column=1, value="Total de Agendamentos:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=len(agendamentos))
        
        total_receita = sum(ag.valor_pago for ag in agendamentos if ag.valor_pago)
        ws.cell(row=4, column=1, value="Total de Receita:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=f"R$ {total_receita:.2f}")
        
        # Receita por profissional
        ws.cell(row=6, column=1, value="Receita por Profissional").font = Font(bold=True, size=14)
        
        prof_receita = {}
        for ag in agendamentos:
            if ag.valor_pago and ag.valor_pago > 0:
                prof_id = ag.profissional_id
                if prof_id not in prof_receita:
                    prof_receita[prof_id] = 0
                prof_receita[prof_id] += ag.valor_pago
        
        if prof_receita:
            ws.cell(row=8, column=1, value="Profissional").font = Font(bold=True)
            ws.cell(row=8, column=2, value="Receita Total").font = Font(bold=True)
            
            for row, (prof_id, receita) in enumerate(prof_receita.items(), 9):
                profissional = next((p for p in profissionais if p.id == prof_id), None)
                ws.cell(row=row, column=1, value=profissional.nome if profissional else f'Profissional {prof_id}')
                ws.cell(row=row, column=2, value=f"R$ {receita:.2f}")
    
    def _create_general_excel(self, wb, agendamentos, pacientes, profissionais, servicos):
        """Cria planilha Excel geral"""
        ws = wb.active
        ws.title = "Geral"
        
        # Estatísticas
        ws.cell(row=1, column=1, value="Relatório Geral da Clínica").font = Font(bold=True, size=16)
        
        stats = [
            ("Total de Pacientes", len(pacientes)),
            ("Total de Profissionais", len(profissionais)),
            ("Total de Serviços", len(servicos)),
            ("Total de Agendamentos", len(agendamentos))
        ]
        
        for row, (label, value) in enumerate(stats, 3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
        
        # Status dos agendamentos
        ws.cell(row=8, column=1, value="Status dos Agendamentos").font = Font(bold=True, size=14)
        
        status_count = {}
        for ag in agendamentos:
            status = ag.status
            status_count[status] = status_count.get(status, 0) + 1
        
        if status_count:
            ws.cell(row=10, column=1, value="Status").font = Font(bold=True)
            ws.cell(row=10, column=2, value="Quantidade").font = Font(bold=True)
            
            for row, (status, count) in enumerate(status_count.items(), 11):
                ws.cell(row=row, column=1, value=status)
                ws.cell(row=row, column=2, value=count)

# Instância global
reports = ReportGenerator()
